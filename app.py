from flask import Flask, render_template, request
import requests
from datetime import datetime, timedelta
from dotenv import load_dotenv
import os
import pycountry
from pycountry_convert import country_alpha3_to_country_alpha2
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file


load_dotenv()


CURRENCYBEACON_API_KEY = os.getenv("CURRENCYBEACON_API_KEY")

app = Flask(__name__)
# ==========================================
# Flag Generator
# ==========================================

def get_currency_flag(currency_code):

    currency_country_map = {

        "USD": "US",
        "DZD": "DZ",
        "EUR": "EU",
        "JPY": "JP",
        "CAD": "CA",
        "GBP": "GB",
        "AUD": "AU",
        "CHF": "CH"

    }


    country_code = currency_country_map.get(currency_code)


    if country_code:

        return f"flags/{country_code}.png"


    return "flags/default.png"

# ==========================================
# Conversion History
# ==========================================

history = []
last_conversion = {}
# ==========================================
# Top Movers (Temporary Data)
# ==========================================

top_movers = [

    {"code": "JPY", "change": 1.82},
    {"code": "CAD", "change": 0.91},
    {"code": "GBP", "change": -1.16},
    {"code": "TRY", "change": -0.84}

]
# ==========================================
# Supported Currencies
# ==========================================

# ==========================================
# Load All Currencies
# ==========================================

import json


with open("currencies.json", "r", encoding="utf-8") as file:

    currency_data = json.load(file)



currencies = {}


for item in currency_data:

    code = item["short_code"]

    currencies[code] = {

        "name": item["name"],

        "flag": get_currency_flag(code)

    }



print(f"✅ Loaded {len(currencies)} currencies")
print(currencies["USD"])
print(currencies["DZD"])
print(currencies["INR"])
# ==========================================
# Load Currencies From JSON
# ==========================================

try:

    with open("currencies.json", "r", encoding="utf-8") as file:

        currency_data = json.load(file)


    currencies = {}
    flag_map = {

    "USD": "flags/usa.png",

    "EUR": "flags/europe.png",

    "DZD": "flags/algeria.png",

    "GBP": "flags/uk.png",

    "JPY": "flags/japan.png",

    "CAD": "flags/canada.png",

    "AUD": "flags/australia.png",

    "CHF": "flags/switzerland.png"

}


    for currency in currency_data:
        for currency in currency_data:

            code = currency["short_code"]

            currencies[code] = {

                "name": currency["name"],

                "flag": flag_map.get(code, "flags/default.png")

            }


    print(f"✅ Loaded {len(currencies)} currencies")

    
except Exception as e:

    print("❌ Failed to load currencies.json")
    print(e)


@app.route("/", methods=["GET", "POST"])
def home():

    global last_conversion

    result = None
    rate = None
    amount = 100

    from_currency = "USD"
    to_currency = "EUR"

    error = None

    chart_labels = []
    chart_rates = []
    trend = None
    trend_percent = None

    if request.method == "POST":

        from_currency = request.form["from_currency"]
        to_currency = request.form["to_currency"]

        action = request.form.get("action")

        if action == "swap":

            from_currency, to_currency = to_currency, from_currency

            return render_template(
                "index.html",
                result=None,
                rate=None,
                amount=100,
                from_currency=from_currency,
                to_currency=to_currency,
                currencies=currencies,
                history=history,
                chart_labels=[],
                chart_rates=[],
                last_update=datetime.now().strftime("%H:%M"),
                error=None,
                top_movers=top_movers,
            )

        try:
            amount = float(request.form["amount"])

        except ValueError:

            error = "Please enter a valid amount."

        if amount <= 0:

            error = "Amount must be greater than zero."

        elif from_currency == to_currency:

            error = "Please choose two different currencies."

        else:

            try:

                # =====================================
                # Current Exchange Rate
                # CurrencyBeacon API
                # =====================================

                url = (
                    f"https://api.currencybeacon.com/v1/latest"
                    f"?api_key={CURRENCYBEACON_API_KEY}"
                    f"&base={from_currency}"
                )

                response = requests.get(url, timeout=5)

                data = response.json()

                if "rates" in data:

                    rate = data["rates"][to_currency]

                    

                    result = amount * rate

                    history.insert(0, {

                        "datetime": datetime.now().strftime("%Y-%m-%d %H:%M"),

                        "amount": amount,

                        "from": from_currency,

                        "to": to_currency,

                        "rate": round(rate, 6),

                        "result": round(result, 2)

                    })
                    last_conversion = {

                        "from": from_currency,

                        "to": to_currency,

                        "rate": round(rate, 6),

                        "trend": trend,

                        "trend_percent": trend_percent

                    }

                    history[:] = history[:10]
                     # =====================================
                    # Last 7 Days Chart
                    # CurrencyBeacon Timeseries
                    # =====================================

                    end_date = datetime.today().date()
                    start_date = end_date - timedelta(days=6)

                    chart_url = (
                        f"https://api.currencybeacon.com/v1/timeseries"
                        f"?api_key={CURRENCYBEACON_API_KEY}"
                        f"&base={from_currency}"
                        f"&symbols={to_currency}"
                        f"&start_date={start_date}"
                        f"&end_date={end_date}"
                    )

                    chart_response = requests.get(chart_url, timeout=5)
                    chart_data = chart_response.json()
                    if "response" in chart_data:

                        chart_labels.clear()
                        chart_rates.clear()

                        for day, values in sorted(chart_data["response"].items()):

                            if to_currency in values:

                                chart_labels.append(day)

                                chart_rates.append(values[to_currency])
                                # =====================================
                                # Price Trend (Yesterday vs Today)
                                # =====================================

                                if len(chart_rates) >= 2:

                                    yesterday = chart_rates[-2]
                                    today = chart_rates[-1]

                                    trend = "up" if today > yesterday else "down"

                                    trend_percent = round(((today - yesterday) / yesterday) * 100, 2)
                                    print("Yesterday:", yesterday)
                                    print("Today:", today)
                                    print("Trend:", trend)
                                    print("Percent:", trend_percent)
            except requests.exceptions.RequestException:

                error = "Connection error. Please try again later."

            except KeyError:

                error = "Currency not supported."

    last_update = datetime.now().strftime("%H:%M")

    return render_template(

        "index.html",

        result=result,

        rate=rate,

        amount=amount,

        from_currency=from_currency,

        to_currency=to_currency,

        currencies=currencies,

        history=history,

        chart_labels=chart_labels,

        chart_rates=chart_rates,

        trend=trend,

        trend_percent=trend_percent,

        last_update=last_update,

        error=error,

        top_movers=top_movers

    )
import csv
from flask import Response


@app.route("/export-history")
def export_history():

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Conversion History"

    # ==========================
    # Title
    # ==========================

    sheet.merge_cells("A1:F1")
    sheet["A1"] = "DENARIUS"
    sheet["A1"].font = Font(size=18, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A2:F2")
    sheet["A2"] = "Global Currency Exchange"
    sheet["A2"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A3:F3")
    sheet["A3"] = "Conversion History Report"
    sheet["A3"].font = Font(bold=True)
    sheet["A3"].alignment = Alignment(horizontal="center")

    sheet["A5"] = "Generated Date:"
    sheet["B5"] = datetime.now().strftime("%Y-%m-%d")

    sheet["D5"] = "Generated Time:"
    sheet["E5"] = datetime.now().strftime("%H:%M")

    # ==========================
    # Headers
    # ==========================
    sheet["A6"] = "Current Pair:"

    if last_conversion:
        sheet["B6"] = f'{last_conversion["from"]} → {last_conversion["to"]}'
    else:
        sheet["B6"] = "-"


    sheet["D6"] = "24H Change:"

    if last_conversion:

        if last_conversion["trend"] == "up":
            sheet["E6"] = f'▲ +{last_conversion["trend_percent"]}%'

        elif last_conversion["trend"] == "down":
            sheet["E6"] = f'▼ -{last_conversion["trend_percent"]}%'

        else:
            sheet["E6"] = "No Change"

    else:
        sheet["E6"] = "-"


    sheet["A7"] = "Current Rate:"

    if last_conversion:
        sheet["B7"] = (
            f'1 {last_conversion["from"]} = '
            f'{last_conversion["rate"]} '
            f'{last_conversion["to"]}'
    )
    else:
        sheet["B7"] = "-"


    sheet["D7"] = "Total Conversions:"
    sheet["E7"] = len(history)
    headers = [
        "Date & Time",
        "From",
        "To",
        "Amount",
        "Rate",
        "Result"
    ]

    header_fill = PatternFill(
        start_color="2563EB",
        end_color="2563EB",
        fill_type="solid"
    )

    row = 7

    for column, header in enumerate(headers, start=1):

        cell = sheet.cell(row=row, column=column)

        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ==========================
    # History
    # ==========================

    row = 8

    for item in history:

        sheet.cell(row=row, column=1).value = item["datetime"]
        sheet.cell(row=row, column=2).value = item["from"]
        sheet.cell(row=row, column=3).value = item["to"]
        sheet.cell(row=row, column=4).value = item["amount"]
        sheet.cell(row=row, column=5).value = item["rate"]
        sheet.cell(row=row, column=6).value = item["result"]

        row += 1

    # ==========================
    # Auto Width
    # ==========================

    for column in sheet.columns:

        length = 0
        letter = get_column_letter(column[0].column)

        for cell in column:

            try:
                if len(str(cell.value)) > length:
                    length = len(str(cell.value))
            except:
                pass

        sheet.column_dimensions[letter].width = length + 4

    # ==========================
    # Save
    # ==========================

    file = BytesIO()
    workbook.save(file)
    file.seek(0)

    return send_file(
        file,
        as_attachment=True,
        download_name="Denarius_History.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":

    app.run(debug=True)
